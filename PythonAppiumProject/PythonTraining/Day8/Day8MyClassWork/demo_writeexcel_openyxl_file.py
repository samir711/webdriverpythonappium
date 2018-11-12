import openpyxl

#  Works with  xlsx format, Use openpyxl for xlsx
#  Reference - https://hackernoon.com/working-with-spreadsheets-using-python-part-2-c8b4ea7e2b65
#  Reference  - https://hackernoon.com/working-with-spreadsheets-using-python-part-1-380a120387f
excel_path = ".\..\..\data\\excel_files\\test.xlsx"

#
# def main():
#     excel_workbook = openpyxl.Workbook()
#     sheet = excel_workbook.create_sheet('TestResult')
#     for rw in range(5):
#         for col in range(5):
#             row_index = rw + 2
#             col_index = col + 2
#             sheet.cell(row=row_index, column=col_index).value = row_index * 10
#     excel_workbook.save(excel_path, )
#
#
# if __name__ == "__main__":
#     main()


def main():
    excel_workbook = openpyxl.load_workbook(excel_path)
    sheet = excel_workbook['TestResult']
    for rw in range(5):
        for col in range(5):
            row_index = rw + 2
            col_index = col + 2
            sheet.cell(row=row_index, column=col_index).value = row_index * 10
    excel_workbook.save(excel_path)


if __name__ == "__main__":
    main()